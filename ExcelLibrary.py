from openpyxl import load_workbook
from datetime import datetime

class ExcelLibrary:
    def load_workbook(self, file_path):
        """Loads an Excel workbook and returns the workbook object."""
        workbook = load_workbook(filename=file_path)
        return workbook

    def get_sheet_names(self, workbook):
        """Returns the names of the sheets in the workbook."""
        return workbook.sheetnames

    def get_cell_value(self, file_path, sheet_name, cell):
        """Returns the value of a specific cell in a specific sheet."""
        workbook = load_workbook(filename=file_path)
        sheet = workbook[sheet_name]
        return sheet[cell].value

    def get_row_values(self, file_path, sheet_name, start_cell, end_cell):
        """Returns the values of a range of cells in a specific row."""
        workbook = load_workbook(filename=file_path)
        sheet = workbook[sheet_name]
        values = []
        for row in sheet[start_cell:end_cell]:
            for cell in row:
                values.append(cell.value)
        return values

    def write_cell_value(self, file_path, sheet_name, cell, value):
        """Writes a value to a specific cell in a specific sheet."""
        workbook = load_workbook(filename=file_path)
        sheet = workbook[sheet_name]
        sheet[cell] = value
        workbook.save(filename=file_path)

    def get_current_time(self):
        """Returns the current time as a string."""
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
